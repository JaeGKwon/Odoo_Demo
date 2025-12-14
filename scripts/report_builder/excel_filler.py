#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 값 채우기 엔진

매핑 파일을 기반으로 Excel 템플릿에 데이터를 채웁니다.
"""

import json
import shutil
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Any

try:
    import openpyxl
    from openpyxl.utils import column_index_from_string
except ImportError:
    print("openpyxl이 필요합니다. 설치하세요: pip install openpyxl")
    exit(1)


class ExcelFiller:
    """Excel 템플릿에 데이터를 채우는 클래스"""
    
    def __init__(self, mapping_path: str, template_dir: str = None):
        """
        Args:
            mapping_path: 매핑 JSON 파일 경로
            template_dir: 템플릿 파일이 있는 디렉토리 (기본: 매핑 파일과 같은 위치)
        """
        self.mapping_path = Path(mapping_path)
        
        if not self.mapping_path.exists():
            raise FileNotFoundError(f"매핑 파일을 찾을 수 없습니다: {mapping_path}")
        
        with open(self.mapping_path, 'r', encoding='utf-8') as f:
            self.mapping = json.load(f)
        
        # 템플릿 경로 결정
        if template_dir:
            self.template_path = Path(template_dir) / self.mapping['template_file']
        else:
            # 프로젝트 루트에서 찾기
            project_root = self.mapping_path.parent.parent
            self.template_path = project_root / self.mapping['template_file']
        
        if not self.template_path.exists():
            raise FileNotFoundError(f"템플릿 파일을 찾을 수 없습니다: {self.template_path}")
        
        # 제품 코드 -> 매핑 정보 인덱스 생성
        self._build_product_index()
        
    def _build_product_index(self):
        """제품 코드로 빠르게 찾기 위한 인덱스를 생성합니다."""
        self.product_index = {}
        
        for product in self.mapping.get('products', []):
            code = product.get('odoo_product_code', '')
            excel_name = product.get('excel_name', '')
            
            if code:
                self.product_index[code] = product
            if excel_name:
                # Excel 이름으로도 찾을 수 있게
                self.product_index[excel_name] = product
                # 정규화된 이름도 추가
                normalized = excel_name.replace(' ', '').upper()
                self.product_index[normalized] = product
    
    def find_product(self, code_or_name: str) -> Optional[dict]:
        """
        제품 코드 또는 이름으로 매핑 정보를 찾습니다.
        
        Args:
            code_or_name: Odoo 제품 코드 또는 Excel 품명
        
        Returns:
            매핑 정보 또는 None
        """
        # 직접 매칭
        if code_or_name in self.product_index:
            return self.product_index[code_or_name]
        
        # 정규화된 이름으로 매칭
        normalized = code_or_name.replace(' ', '').upper()
        if normalized in self.product_index:
            return self.product_index[normalized]
        
        # 부분 매칭 시도
        for key, product in self.product_index.items():
            if code_or_name.upper() in key.upper() or key.upper() in code_or_name.upper():
                return product
        
        return None
    
    def _parse_cell_ref(self, cell_ref: str) -> tuple:
        """
        셀 참조(예: 'E6')를 (열, 행) 튜플로 변환합니다.
        
        Returns:
            (column_index, row_number) - 1-based
        """
        import re
        match = re.match(r'^([A-Z]+)(\d+)$', cell_ref.upper())
        if not match:
            raise ValueError(f"잘못된 셀 참조: {cell_ref}")
        
        col_letter = match.group(1)
        row_num = int(match.group(2))
        col_idx = column_index_from_string(col_letter)
        
        return col_idx, row_num
    
    def create_report(self, 
                      output_path: str,
                      date: str = None,
                      person: str = None,
                      stock_data: Dict[str, dict] = None) -> str:
        """
        리포트를 생성합니다.
        
        Args:
            output_path: 출력 파일 경로
            date: 리포트 날짜 (기본: 오늘)
            person: 담당자 이름
            stock_data: 재고 데이터 딕셔너리
                {
                    '제품코드': {
                        'incoming': 100.0,  # 입고
                        'outgoing': 50.0,   # 출고
                        'unit': 'L'         # 단위 (선택)
                    }
                }
        
        Returns:
            생성된 파일 경로
        """
        # 템플릿 복사
        output = Path(output_path)
        output.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(self.template_path, output)
        
        # 워크북 열기
        wb = openpyxl.load_workbook(str(output))
        ws = wb.active
        
        # 날짜 채우기
        if date:
            date_cell = self.mapping.get('date_cell', 'A3')
            col, row = self._parse_cell_ref(date_cell)
            ws.cell(row=row, column=col, value=date)
        
        # 담당자 채우기
        if person:
            person_cell = self.mapping.get('person_cell', 'J3')
            col, row = self._parse_cell_ref(person_cell)
            ws.cell(row=row, column=col, value=f"담당 : {person}")
        
        # 재고 데이터 채우기
        if stock_data:
            filled_count = 0
            not_found = []
            
            for product_key, data in stock_data.items():
                product_mapping = self.find_product(product_key)
                
                if not product_mapping:
                    not_found.append(product_key)
                    continue
                
                cells = product_mapping.get('cells', {})
                unit = data.get('unit', 'L')  # 기본 단위
                
                # 해당 단위의 셀 찾기
                if unit in cells:
                    cell_info = cells[unit]
                elif cells:
                    # 단위가 없으면 첫 번째 단위 사용
                    cell_info = list(cells.values())[0]
                else:
                    continue
                
                # 입고 값 채우기
                if 'incoming' in data and data['incoming']:
                    try:
                        col, row = self._parse_cell_ref(cell_info['incoming'])
                        ws.cell(row=row, column=col, value=data['incoming'])
                        filled_count += 1
                    except Exception as e:
                        print(f"  입고 셀 오류 ({product_key}): {e}")
                
                # 출고 값 채우기
                if 'outgoing' in data and data['outgoing']:
                    try:
                        col, row = self._parse_cell_ref(cell_info['outgoing'])
                        ws.cell(row=row, column=col, value=data['outgoing'])
                        filled_count += 1
                    except Exception as e:
                        print(f"  출고 셀 오류 ({product_key}): {e}")
            
            print(f"✓ 데이터 채우기 완료: {filled_count}개 셀")
            if not_found:
                print(f"  ⚠ 매핑되지 않은 제품: {len(not_found)}개")
                if len(not_found) <= 5:
                    for nf in not_found:
                        print(f"    - {nf}")
        
        # 저장
        wb.save(str(output))
        print(f"✓ 리포트 저장: {output}")
        
        return str(output)
    
    def fill_cell(self, worksheet, cell_ref: str, value: Any) -> bool:
        """
        특정 셀에 값을 채웁니다.
        
        Args:
            worksheet: openpyxl 워크시트
            cell_ref: 셀 참조 (예: 'E6')
            value: 채울 값
        
        Returns:
            성공 여부
        """
        try:
            col, row = self._parse_cell_ref(cell_ref)
            worksheet.cell(row=row, column=col, value=value)
            return True
        except Exception as e:
            print(f"셀 입력 오류 ({cell_ref}): {e}")
            return False
    
    def get_mapping_summary(self) -> dict:
        """매핑 요약 정보를 반환합니다."""
        products = self.mapping.get('products', [])
        
        total_cells = 0
        for p in products:
            for unit_cells in p.get('cells', {}).values():
                total_cells += len(unit_cells)
        
        return {
            'template_file': self.mapping.get('template_file'),
            'total_products': len(products),
            'total_cells': total_cells,
            'date_cell': self.mapping.get('date_cell'),
            'person_cell': self.mapping.get('person_cell')
        }


def main():
    """테스트 실행"""
    import argparse
    
    parser = argparse.ArgumentParser(description='Excel 템플릿에 데이터 채우기')
    parser.add_argument('mapping', help='매핑 JSON 파일 경로')
    parser.add_argument('-o', '--output', help='출력 파일 경로')
    parser.add_argument('--date', help='리포트 날짜')
    parser.add_argument('--person', help='담당자 이름')
    parser.add_argument('--test', action='store_true', help='테스트 데이터로 채우기')
    parser.add_argument('--info', action='store_true', help='매핑 정보만 출력')
    
    args = parser.parse_args()
    
    try:
        filler = ExcelFiller(args.mapping)
        
        if args.info:
            summary = filler.get_mapping_summary()
            print("\n" + "=" * 50)
            print("매핑 정보")
            print("=" * 50)
            for key, value in summary.items():
                print(f"  {key}: {value}")
            return 0
        
        if args.test:
            # 테스트 데이터
            test_data = {
                'ZP-8': {'incoming': 100, 'outgoing': 50, 'unit': 'L'},
                'ZP-10': {'incoming': 200, 'outgoing': 100, 'unit': 'L'},
                'HITRANS-102K': {'incoming': 150, 'outgoing': 75, 'unit': 'L'},
            }
            
            output = args.output or 'output/test_report.xlsx'
            date = args.date or datetime.now().strftime('%Y-%m-%d')
            person = args.person or '테스트'
            
            filler.create_report(
                output_path=output,
                date=date,
                person=person,
                stock_data=test_data
            )
        else:
            print("--test 옵션으로 테스트하거나, report_generator.py를 사용하세요.")
        
    except Exception as e:
        print(f"오류: {e}")
        return 1
    
    return 0


if __name__ == "__main__":
    exit(main())

