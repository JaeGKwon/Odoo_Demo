#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
매핑 파일 생성기

Excel 템플릿에서 제품명과 셀 위치를 자동으로 추출하여
JSON 매핑 파일을 생성합니다.
"""

import json
import os
import re
from pathlib import Path

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl이 필요합니다. 설치하세요: pip install openpyxl")
    exit(1)


class MappingGenerator:
    """Excel 템플릿에서 제품 매핑을 생성하는 클래스"""
    
    def __init__(self, template_path: str):
        """
        Args:
            template_path: Excel 템플릿 파일 경로
        """
        self.template_path = Path(template_path)
        if not self.template_path.exists():
            raise FileNotFoundError(f"템플릿 파일을 찾을 수 없습니다: {template_path}")
        
        self.workbook = openpyxl.load_workbook(str(self.template_path))
        self.worksheet = self.workbook.active
        self.products = []
        
    def _clean_text(self, value) -> str:
        """셀 값을 정리합니다."""
        if value is None:
            return ""
        return str(value).replace('\n', ' ').strip()
    
    def _extract_products_from_column(self, name_col: int, unit_col: int, 
                                       incoming_col: int, outgoing_col: int,
                                       start_row: int = 6, end_row: int = None) -> list:
        """
        특정 열에서 제품 정보를 추출합니다.
        
        Args:
            name_col: 품명 열 번호 (1-based)
            unit_col: 단위 열 번호
            incoming_col: 입고 열 번호
            outgoing_col: 출고 열 번호
            start_row: 시작 행
            end_row: 종료 행 (None이면 max_row까지)
        """
        if end_row is None:
            end_row = self.worksheet.max_row
            
        products = []
        current_product = None
        
        for row in range(start_row, end_row + 1):
            name_cell = self.worksheet.cell(row=row, column=name_col).value
            unit_cell = self.worksheet.cell(row=row, column=unit_col).value
            
            name = self._clean_text(name_cell)
            unit = self._clean_text(unit_cell)
            
            # 새 제품 시작 (품명이 있는 경우)
            if name:
                if current_product:
                    products.append(current_product)
                
                current_product = {
                    "excel_name": name,
                    "odoo_product_code": self._generate_odoo_code(name),
                    "cells": {}
                }
            
            # 단위가 있으면 셀 정보 추가
            if current_product and unit:
                col_letter_in = get_column_letter(incoming_col)
                col_letter_out = get_column_letter(outgoing_col)
                
                current_product["cells"][unit] = {
                    "incoming": f"{col_letter_in}{row}",
                    "outgoing": f"{col_letter_out}{row}"
                }
        
        # 마지막 제품 추가
        if current_product:
            products.append(current_product)
            
        return products
    
    def _generate_odoo_code(self, excel_name: str) -> str:
        """
        Excel 품명에서 Odoo 제품 코드를 생성합니다.
        (실제로는 수동 매핑이 필요할 수 있음)
        """
        # 간단한 정규화: 특수문자 제거, 공백을 하이픈으로
        code = re.sub(r'[^\w\s-]', '', excel_name)
        code = re.sub(r'\s+', '-', code).upper()
        return code
    
    def extract_all_products(self) -> dict:
        """
        템플릿에서 모든 제품을 추출합니다.
        
        Returns:
            매핑 데이터 딕셔너리
        """
        # 좌측 제품 (B열 품명, C열 단위, E열 입고, F열 출고)
        left_products = self._extract_products_from_column(
            name_col=2, unit_col=3, incoming_col=5, outgoing_col=6
        )
        
        # 우측 제품 (K열 품명, L열 단위, N열 입고, O열 출고)
        right_products = self._extract_products_from_column(
            name_col=11, unit_col=12, incoming_col=14, outgoing_col=15
        )
        
        all_products = left_products + right_products
        
        mapping = {
            "template_file": self.template_path.name,
            "sheet_name": self.worksheet.title,
            "date_cell": "A3",
            "person_cell": "J3",
            "header_info": {
                "title_cell": "A1",
                "approval_cells": {
                    "담당": "O1",
                    "검토": "P1",
                    "승인": "R1"
                }
            },
            "columns": {
                "left": {
                    "category": "A",
                    "product_name": "B",
                    "unit": "C",
                    "prev_stock": "D",
                    "incoming": "E",
                    "outgoing": "F",
                    "current_stock": "G",
                    "monthly_incoming": "H",
                    "monthly_outgoing": "I"
                },
                "right": {
                    "category": "J",
                    "product_name": "K",
                    "unit": "L",
                    "prev_stock": "M",
                    "incoming": "N",
                    "outgoing": "O",
                    "current_stock": "P",
                    "monthly_incoming": "Q",
                    "monthly_outgoing": "R"
                }
            },
            "products": all_products
        }
        
        return mapping
    
    def save_mapping(self, output_path: str) -> None:
        """
        매핑 데이터를 JSON 파일로 저장합니다.
        
        Args:
            output_path: 출력 파일 경로
        """
        mapping = self.extract_all_products()
        
        output = Path(output_path)
        output.parent.mkdir(parents=True, exist_ok=True)
        
        with open(output, 'w', encoding='utf-8') as f:
            json.dump(mapping, f, ensure_ascii=False, indent=2)
        
        print(f"✓ 매핑 파일 저장 완료: {output}")
        print(f"  - 총 제품 수: {len(mapping['products'])}")
        
    def print_summary(self) -> None:
        """매핑 요약을 출력합니다."""
        mapping = self.extract_all_products()
        
        print("\n" + "=" * 60)
        print("제품 매핑 요약")
        print("=" * 60)
        print(f"템플릿 파일: {mapping['template_file']}")
        print(f"시트명: {mapping['sheet_name']}")
        print(f"날짜 셀: {mapping['date_cell']}")
        print(f"담당자 셀: {mapping['person_cell']}")
        print(f"총 제품 수: {len(mapping['products'])}")
        print()
        
        print("처음 10개 제품:")
        for i, prod in enumerate(mapping['products'][:10], 1):
            print(f"  {i}. {prod['excel_name']}")
            print(f"     Odoo 코드: {prod['odoo_product_code']}")
            print(f"     단위별 셀: {list(prod['cells'].keys())}")


def main():
    """메인 실행 함수"""
    import argparse
    
    parser = argparse.ArgumentParser(description='Excel 템플릿에서 제품 매핑 파일 생성')
    parser.add_argument('template', help='Excel 템플릿 파일 경로')
    parser.add_argument('-o', '--output', default='mappings/product_mapping.json',
                        help='출력 매핑 파일 경로 (기본: mappings/product_mapping.json)')
    parser.add_argument('--summary', action='store_true', help='요약만 출력')
    
    args = parser.parse_args()
    
    try:
        generator = MappingGenerator(args.template)
        
        if args.summary:
            generator.print_summary()
        else:
            generator.save_mapping(args.output)
            generator.print_summary()
            
    except Exception as e:
        print(f"오류: {e}")
        return 1
    
    return 0


if __name__ == "__main__":
    exit(main())

